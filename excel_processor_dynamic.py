import pandas as pd
import time
import psycopg2
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import tkinter.messagebox as messagebox  # если не импортировано
import requests
import helpers
from openpyxl.styles import Alignment



def get_id_with_leading_zeros(value):
    if pd.isna(value):
        return None
    str_value = str(value).strip()
    if 'E+' in str_value.upper() or 'e+' in str_value:
        try:
            float_val = float(str_value)
            int_val = int(float_val)
            str_value = str(int_val)
        except:
            pass
    if '.' in str_value:
        try:
            float_val = float(str_value)
            if float_val.is_integer():
                str_value = str(int(float_val))
        except:
            pass
    return str_value


def fetch_html_with_timeout(url, timeout=5):
    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Ошибка загрузки HTML: {e}")
        return None


def analyze_application_from_html(html_content, return_deadline=False):
    try:
        soup = BeautifulSoup(html_content, "html.parser")
        result = helpers.analyzeFullApplication(soup, return_deadline=return_deadline)
        return result  # будет кортеж или строка — в зависимости от return_deadline
    except Exception as e:
        return (f"Ошибка анализа HTML: {e}", None) if return_deadline else f"Ошибка анализа HTML: {e}"


def query_postgres_by_app_id(app_id, conn):
    try:
        print(f"🔍 Выполняется SQL-запрос для app_id: {app_id}")
        with conn.cursor() as cur:
            cur.execute("""
                SELECT i.id, i.requestnumber, h.status AS status_name, h.creation_date
                FROM history.application_info i
                LEFT JOIN history.status_history h ON h.applicationinfo_id = i.id
                LEFT JOIN history.status_go s ON s.id = h.statusgo_id
                WHERE i.requestnumber  = %s
                ORDER BY h.creation_date DESC;
            """, (app_id,))

            results = cur.fetchall()  # Получаем все строки
            print(f"✅ Получено {len(results)} строк из БД")

            if results:
                first_result = results[0]
                status = first_result[2] or "Статус отсутствует"

                # Если текущий статус — оплата, ищем предыдущий не-оплатный статус
                if status in ["PAYED", "WAITING_FOR_PAYMENT"]:
                    index = 1  # начинаем с элемента после first_result
                    while index < len(results):
                        temp_status = results[index][2]
                        if temp_status not in ["PAYED", "WAITING_FOR_PAYMENT"]:
                            status = temp_status or "Статус отсутствует"
                            break
                        index += 1
                    else:
                        # если все статусы — оплатные
                        status = "Статус перед оплатой не найден"

            #     print("Статус:", status)
            # else:
            #     status = "Нет результатов"


                # TODO: обработать PAYED WAITING FOR PAYMENT
                date = first_result[3].strftime('%Y-%m-%d %H:%M:%S.%f') if first_result[3] else "Дата отсутствует"

                # Анализируем статус
                dbConclusion = helpers.analyzeDBStatuses(status)

                # Если тебе нужно вернуть все результаты — можно так:
                # return dbConclusion, date, results
                if helpers.hasTechErrors(results):
                    return dbConclusion + ". Есть TECH_ERROR", date
                return dbConclusion, date
            else:
                return "Нет истории по номеру заявления", ""

    except Exception as e:
        return f"Ошибка запроса: {e}", ""



def preserve_excel_formatting(original_file, output_file, df_updated, sheet_index):
    try:
        wb = load_workbook(original_file)
        ws = wb.worksheets[sheet_index]

        # Находим индекс столбца с комментариями
        comment_col_index = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and 'Комментарий АО НИТ' in str(cell.value):
                comment_col_index = col_idx
                break
        if comment_col_index is None:
            return False

        # Создаём объект с wrapText
        wrap_alignment = Alignment(wrap_text=True)

        for row_idx in range(2, len(df_updated) + 2):
            df_row_idx = row_idx - 2
            comment_value = df_updated.iloc[df_row_idx]['Комментарий АО НИТ']

            if pd.notna(comment_value) and comment_value != '':
                cell = ws.cell(row=row_idx, column=comment_col_index)
                cell.value = comment_value
                cell.alignment = wrap_alignment

                # ✅ ВАЖНО: включаем автоматическую высоту строки
                ws.row_dimensions[row_idx].height = None

        wb.save(output_file)
        return True

    except Exception as e:
        print(f"Ошибка сохранения Excel: {e}")
        return False


def process_html_sheet(file_path, output_path):
    try:
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', dtype=str)
    except Exception as e:
        print(f"Ошибка чтения первой страницы Excel: {e}")
        return

    identifier_col = comment_col = None
    for col in df.columns:
        if 'Идентификатор' in str(col):
            identifier_col = col
        elif 'Комментарий АО НИТ' in str(col):
            comment_col = col

    if identifier_col is None:
        print("Столбец идентификатора не найден на первой странице")
        return

    if comment_col is None:
        comment_col = 'Комментарий АО НИТ'
        df[comment_col] = ''

    print(f"Найден столбец идентификатора: {identifier_col}")
    print(f"Найден столбец комментариев: {comment_col}")
    print(f"Всего строк для обработки: {len(df)}")

    base_url = "http://192.168.130.100/csp/iiscon/isc.util.About.cls?Action=4&appId="
    successful_count = 0
    failed_count = 0

    for index, row in df.iterrows():
        app_id_raw = row[identifier_col]
        if pd.isna(app_id_raw) or str(app_id_raw).strip() in ('', 'nan'):
            continue

        app_id = get_id_with_leading_zeros(app_id_raw)
        if app_id is None:
            continue

        url = base_url + app_id
        print(f"\n[{index + 1}/{len(df)}] Обработка заявки ID: {app_id}")

        html_content = fetch_html_with_timeout(url, timeout=5)
        if html_content:
            conclusion = analyze_application_from_html(html_content, return_deadline=False)
            df.at[index, comment_col] = conclusion
            successful_count += 1
            print(f"✓ Успешно: {conclusion}")
        else:
            df.at[index, comment_col] = "Ошибка: не удалось получить данные"
            failed_count += 1
            print("✗ Не удалось получить данные")

        time.sleep(0.5)

    preserved = preserve_excel_formatting(file_path, output_path, df, sheet_index=0)
    if not preserved:
        print("⚠️ Форматирование не сохранено, сохраняем обычным способом")
        df.to_excel(output_path, index=False, engine='openpyxl')

    print(f"\n=== РЕЗУЛЬТАТЫ ===")
    print(f"Успешно обработано: {successful_count}")
    print(f"Ошибок: {failed_count}")
    print(f"Результаты сохранены в: {output_path}")


def process_pep_sheet_with_full_analysis(file_path, output_path, db_config):
    try:
        df = pd.read_excel(file_path, sheet_name=2, engine='openpyxl', dtype=str)
    except Exception as e:
        print(f"Ошибка чтения листа 'ПЭП': {e}")
        return

    identifier_col = comment_col = pep_col = None
    for col in df.columns:
        if 'Идентификатор заявки' in str(col):
            identifier_col = col
        elif 'Комментарий АО НИТ' in str(col):
            comment_col = col
        elif 'Номер заявления' in str(col):
            pep_col = col

    if identifier_col is None or pep_col is None:
        print("Не найдены нужные столбцы на листе 'ПЭП'")
        return

    if comment_col is None:
        comment_col = 'Комментарий АО НИТ'
        df[comment_col] = ''

    print(f"Найден столбец идентификатора: {identifier_col}")
    print(f"Найден столбец PEP: {pep_col}")
    print(f"Найден столбец комментариев: {comment_col}")
    print(f"Всего строк для обработки: {len(df)}")

    successful_count = 0
    failed_count = 0
    BATCH_SIZE = 100
    conn = None

    for index, row in df.iterrows():
        app_id = row[identifier_col]
        pep_id = row[pep_col]
        if pd.isna(app_id) or pd.isna(pep_id):
            continue

        if index % BATCH_SIZE == 0:
            if conn:
                conn.close()
            try:
                conn = psycopg2.connect(
                    host=db_config["host"],
                    port=db_config["port"],
                    dbname=db_config["dbname"],
                    user=db_config["user"],
                    password=db_config["password"],
                    sslmode="disable",
                    connect_timeout=5,
                    options='-c statement_timeout=5000'
                )
                print(f"🔌 Подключение к БД установлено (batch {index // BATCH_SIZE + 1})")
            except Exception as e:
                print(f"❌ Ошибка подключения к БД: {e}")
                conn = None

        print(f"\n[{index + 1}/{len(df)}] Обработка PEP-записи: ID={app_id}, Номер={pep_id}")

        html_content = fetch_html_with_timeout(
            f"http://192.168.130.100/csp/iiscon/isc.util.About.cls?Action=4&appId={app_id}")

        if html_content:
            html_conclusion, deadline = analyze_application_from_html(html_content, return_deadline=True)
        else:
            html_conclusion, deadline = "Ошибка загрузки HTML", None
            df.at[index, comment_col] = html_conclusion
            print(f"✓ Ошибка: {html_conclusion}")
            continue

        # ✅ Если сразу нужно остановить и перейти к следующей записи
        if html_conclusion in [
            "ГУ оказана несвоевременно.",
            "ГУ оказана несвоевременно. Рассмотреть на стороне ГО."
        ]:
            df.at[index, comment_col] = html_conclusion
            print(f"✓ Успешно: {html_conclusion}")
            successful_count += 1
            time.sleep(0.5)
            continue  # ⬅️ Переход на следующую строку DataFrame

        if not str(pep_id).startswith(('0', '1')):
            df.at[index, comment_col] = html_conclusion
            print(f"✓ Успешно: {html_conclusion}")
            successful_count += 1
            time.sleep(0.5)
            continue  # ⬅️ Переход на следующую строку DataFrame

        if not conn or conn.closed != 0:
            df.at[index, comment_col] = "Ошибка подключения к БД"
            failed_count += 1
            print("⛔ Пропуск из-за отсутствия соединения")
            continue

        db_status, finishDate = query_postgres_by_app_id(pep_id, conn)

        if db_status == html_conclusion:
            df.at[index, comment_col] = html_conclusion
            print(f"✓ Успешно: {html_conclusion}")

        elif db_status == "Рассмотреть на SHEP" and html_conclusion in ["ГУ на исполнении.", "ГУ на исполнении. Рассмотреть на стороне ГО."]:
            conc = "ГУ на исполнении. Рассмотреть на стороне ГО."
            df.at[index, comment_col] = conc
            print(conc)
        elif db_status in ["FINISHED", "CANCELLED", "APPROVED"]:
            print(f"db date {finishDate}")
            print(f"shina deadline {deadline}")
            if helpers.checkStatusDeadline(finishDate, deadline):
                conc = "ГУ оказана своевременно."
                if html_conclusion != "ГУ оказана несвоевременно." or "ГУ оказана несвоевременно. Рассмотреть на стороне ГО.":
                    conc += " Однако статус исполнения отсутствует в ИИС ЦОН."
                df.at[index, comment_col] = conc
                print(conc)
            else:
                conc = "ГУ оказана несвоевременно. Рассмотреть на стороне ГО."
                if html_conclusion != "ГУ оказана несвоевременно." or "ГУ оказана несвоевременно. Рассмотреть на стороне ГО.":
                    conc += " Однако статус исполнения отсутствует в ИИС ЦОН."
                df.at[index, comment_col] = conc
                print(conc)
        else:
            df.at[index, comment_col] = f"HTML: {html_conclusion} | БД: {db_status}"
            print(f"HTML: {html_conclusion} | БД: {db_status}")

        successful_count += 1
        time.sleep(0.5)

    preserved = preserve_excel_formatting(file_path, output_path, df, sheet_index=2)
    if not preserved:
        print("⚠️ Форматирование не сохранено, сохраняем обычным способом")
        df.to_excel(output_path, index=False, engine='openpyxl')

    print(f"\n=== РЕЗУЛЬТАТЫ (ПЭП) ===")
    print(f"Успешно обработано: {successful_count}")
    print(f"Ошибок: {len(df) - successful_count}")
    print(f"Результаты сохранены в: {output_path}")


def process_combined_excel_pipeline(file_path, output_path, db_config):
    process_html_sheet(file_path, output_path)

    # ✅ Подтверждение перед переходом к ПЭП
    proceed = messagebox.askyesno(
        "Переход к ПЭП-листу",
        "HTML-лист успешно обработан.\nХотите продолжить обработку ПЭП-листа с подключением к базе данных?"
    )

    if not proceed:
        print("⏹️ Пользователь завершил обработку на этапе HTML.")
        return  # Просто выходим — HTML уже обработан

    try:
        # 🔄 Передаём db_config напрямую — подключение будет открыто внутри
        process_pep_sheet_with_full_analysis(file_path, output_path, db_config)
    except Exception as e:
        print(f"❌ Ошибка во время анализа ПЭП: {e}")


